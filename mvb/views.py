from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Q
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login, logout
from django.views.decorators.http import require_GET
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib import messages
from decimal import Decimal
from calendar import monthrange
from django.db import transaction
from .models import Cliente, Presenca, BonusPayment
from .models import (
        LavagemCarreta, LavadorSujoEntry, LavadorCargaEntry, Cliente
)
from .forms import ClienteForm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import pandas as pd
from django.utils import timezone
from datetime import datetime, timedelta,date
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import matplotlib.pyplot as plt
import base64
from weasyprint import HTML
from django.core.paginator import Paginator
from .models import (
    Funcionario, Funcao, Financeiro,
    LavagemCarreta, LavadorSujoEntry, LavadorCargaEntry, TipoCaixa, TipoProduto, Profile
)
from .forms import (
    FuncionarioForm, FuncaoForm, FinanceiroForm,
    LavagemCarretaForm, LavadorSujoForm, LavadorCargaForm,
    TipoCaixaForm, TipoProdutoForm, UserRegisterForm
)
from .permissions import admin_required, entry_allowed


# Registro
def register_request(request):
    if request.method == "POST":
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = True
            user.first_name = form.cleaned_data.get('first_name') or ''
            user.save()
            user.profile.is_approved = False
            user.profile.save()
            messages.success(request, "Solicitação enviada. Aguarde aprovação do administrador.")
            return render(request, 'mvb/register_done.html', {'user': user})
    else:
        form = UserRegisterForm()
    return render(request, 'mvb/register.html', {'form': form})

@admin_required
def lista_pedidos_pendentes(request):
    pendentes = Profile.objects.filter(is_approved=False).select_related('user')
    return render(request, 'mvb/pedidos_pendentes.html', {'pendentes': pendentes})

@admin_required
def aprovar_usuario(request, user_id):
    u = get_object_or_404(User, pk=user_id)
    u.profile.is_approved = True
    u.profile.save()
    messages.success(request, f"Usuário {u.username} aprovado.")
    return redirect('lista_pedidos_pendentes')

@require_GET
def logout_view(request):
    """Efetua logout e redireciona para a tela de login."""
    logout(request)
    messages.success(request, "Você saiu do sistema com sucesso.")
    return redirect('login')

# Dashboard
@login_required
def mvb_dashboard(request):
    hoje = date.today()
    funcionarios = Funcionario.objects.filter(ativo=True)
    inicio_sem = hoje - timedelta(days=hoje.weekday())
    fim_sem = inicio_sem + timedelta(days=6)

    elegiveis_semana = [
        f for f in funcionarios if Presenca.objects.filter(
            funcionario=f,
            data__range=(inicio_sem, fim_sem),
            status="F"
        ).count() == 0
    ]

    elegiveis_mes = [
        f for f in funcionarios if Presenca.objects.filter(
            funcionario=f,
            data__month=hoje.month,
            data__year=hoje.year,
            status="F"
        ).count() == 0
    ]

    receita_carretas = LavagemCarreta.objects.filter(data__year=hoje.year, data__month=hoje.month).aggregate(
        total=Sum(ExpressionWrapper(F('quantidade_caixas') * F('valor_por_caixa'), output_field=DecimalField()))
    )['total'] or Decimal('0.00')

    receita_sujo = LavadorSujoEntry.objects.filter(data__year=hoje.year, data__month=hoje.month).aggregate(
        total=Sum(ExpressionWrapper(F('quantidade_caixas') * F('valor_por_caixa'), output_field=DecimalField()))
    )['total'] or Decimal('0.00')

    receita_carga = LavadorCargaEntry.objects.filter(data__year=hoje.year, data__month=hoje.month).aggregate(
        total=Sum('valor_rendido')
    )['total'] or Decimal('0.00')

    receita_total = float(receita_carretas + receita_sujo + receita_carga)
    financeiro = Financeiro.objects.filter(ano=hoje.year, mes=hoje.month).first()
    despesas_total = float(financeiro.total) if financeiro else 0.0

    return render(request, 'mvb/dashboard.html', {
        'receita_total': receita_total,
        'despesas_total': despesas_total,
        'financeiro': financeiro,
        'qtd_elegiveis_semana': len(elegiveis_semana),
        'qtd_elegiveis_mes': len(elegiveis_mes),
    })

# CRUD simplificados (use decorators conforme necessidade)
@login_required
def lista_funcoes(request):
    funcoes = Funcao.objects.all()
    return render(request, 'mvb/lista_funcoes.html', {'funcoes': funcoes})

@admin_required
def nova_funcao(request):
    form = FuncaoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Função criada.")
        return redirect('lista_funcoes')
    return render(request, 'mvb/form.html', {'form': form, 'titulo': 'Nova Função'})

@login_required
def lista_funcionarios(request):
    funcionarios = Funcionario.objects.select_related('funcao').all()
    return render(request, 'mvb/lista_funcionarios.html', {'funcionarios': funcionarios})

# permite criar funcionário para quem tem permissão mvb.add_funcionario
@login_required
@permission_required('mvb.add_funcionario', raise_exception=True)
def novo_funcionario(request):
    form = FuncionarioForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Funcionário criado.")
            return redirect('lista_funcionarios')
        else:
            messages.error(request, "Erros no formulário. Verifique os campos.")
    return render(request, 'mvb/form.html', {'form': form, 'titulo': 'Novo Funcionário'})

@admin_required
def editar_funcionario(request, pk):
    obj = get_object_or_404(Funcionario, pk=pk)
    form = FuncionarioForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, "Funcionário atualizado.")
        return redirect('lista_funcionarios')
    return render(request, 'mvb/form.html', {'form': form, 'titulo': 'Editar Funcionário'})

@admin_required
def deletar_funcionario(request, pk):
    obj = get_object_or_404(Funcionario, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, "Funcionário excluído.")
        return redirect('lista_funcionarios')
    return render(request, 'mvb/confirmar_exclusao.html', {'obj': obj})

#Cadastrar caixas
def lista_tipos_caixa(request):
    tipos = TipoCaixa.objects.filter(ativo=True)
    return render(request, 'mvb/lista_tipos_caixa.html', {'tipos': tipos})

def criar_tipo_caixa(request):
    if request.method == 'POST':
        form = TipoCaixaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_tipos_caixa')
    else:
        form = TipoCaixaForm()

    return render(request, 'mvb/criar_tipo_caixa.html', {'form': form})

def excluir_tipo_caixa(request, pk):
    tipo = get_object_or_404(TipoCaixa, pk=pk)

    if request.method == 'POST':
        tipo.ativo = False
        tipo.save()
        return redirect('lista_tipos_caixa')

    return render(request, 'mvb/confirmar_excluir_caixa.html', {'tipo': tipo})

#Cadastrar produto
def lista_tipos_produto(request):
    tipos = TipoProduto.objects.filter(ativo=True)
    return render(request, 'mvb/lista_tipos_produto.html', {'tipos': tipos})

def criar_tipo_produto(request):
    if request.method == 'POST':
        form = TipoProdutoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_tipos_produto')
    else:
        form = TipoProdutoForm()

    return render(request, 'mvb/criar_tipo_produto.html', {'form': form})

def excluir_tipo_produto(request, pk):
    tipo = get_object_or_404(TipoProduto, pk=pk)

    if request.method == 'POST':
        tipo.ativo = False
        tipo.save()
        return redirect('lista_tipos_produto')

    return render(request, 'mvb/confirmar_excluir_produto.html', {'tipo': tipo})

@login_required
def lista_lavagens(request):
    lavagens_list = LavagemCarreta.objects.select_related('tipo_caixa','tipo_produto').all().order_by('-data')

    paginator = Paginator(lavagens_list, 6)  # 10 por página
    page_number = request.GET.get('page')
    lavagens = paginator.get_page(page_number)

    return render(request, 'mvb/lista_lavagens.html', {'lavagens': lavagens})

@admin_required
def editar_lavagem(request, pk):
    lavagem = get_object_or_404(LavagemCarreta, pk=pk)

    if request.method == "POST":
        form = LavagemCarretaForm(request.POST, instance=lavagem)
        if form.is_valid():
            form.save()
            return redirect("lista_lavagens")
    else:
        form = LavagemCarretaForm(instance=lavagem)

    return render(request, "mvb/editar_lavagem.html", {
        "form": form,
        "lavagem": lavagem
    })

@admin_required
def excluir_lavagem(request, pk):
    lavagem = get_object_or_404(LavagemCarreta, pk=pk)

    if request.method == "POST":
        lavagem.delete()
        return redirect("lista_lavagens")

    return render(request, "mvb/excluir_lavagem.html", {
        "lavagem": lavagem
    })

@entry_allowed
def nova_lavagem(request):
    if request.method == "POST":
        form = LavagemCarretaForm(request.POST)
        if form.is_valid():
            lavagem = form.save(commit=False)
            # atribui o usuário que criou
            lavagem.criado_por = request.user
            lavagem.save()
            messages.success(request, "Lavagem (carreta) criada com sucesso.")
            return redirect("lista_lavagens")  # ajuste para sua URL de lista
        else:
            # mostra erros no template
            messages.error(request, "Erros no formulário. Verifique os campos.")
    else:
        form = LavagemCarretaForm()
    return render(request, "mvb/lavagemcarreta_form.html", {"form": form})

@login_required
def lista_lavador_sujo(request):
    entries = LavadorSujoEntry.objects.select_related('tipo_produto').all().order_by('-data')
    return render(request, 'mvb/lista_lavador_sujo.html', {'entries': entries})

@entry_allowed
def novo_lavador_sujo(request):
    if request.method == "POST":
        form = LavadorSujoForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.criado_por = request.user
            entry.save()
            messages.success(request, "Registro sujo criado com sucesso.")
            return redirect("lista_lavagens")
        else:
            messages.error(request, "Erros no formulário. Verifique os campos.")
    else:
        form = LavadorSujoForm()
    return render(request, "mvb/lavadorsujo_form.html", {"form": form})

@login_required
def lista_lavador_carga(request):
    entries = LavadorCargaEntry.objects.all().order_by('-data')
    return render(request, 'mvb/lista_lavador_carga.html', {'entries': entries})

@entry_allowed
def novo_lavador_carga(request):
    if request.method == "POST":
        form = LavadorCargaForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.criado_por = request.user
            # opcional: calcular quantidade total para consistência
            # entry.quantidade_caixas = entry.total_caixas_por_categoria()
            entry.save()
            messages.success(request, "Registro de carga criado com sucesso.")
            return redirect("lista_lavagens")
        else:
            messages.error(request, "Erros no formulário. Verifique os campos.")
    else:
        form = LavadorCargaForm()
    return render(request, "mvb/lavadorcarga_form.html", {"form": form})

@login_required
def lista_financeiro(request):
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')

    qs = Financeiro.objects.all().order_by('-data')

    if ano:
        qs = qs.filter(ano=int(ano))
    if mes:
        qs = qs.filter(mes=int(mes))

    # 🔹 Totais do mês (para os cards)
    receita_total = qs.aggregate(total=Sum('total'))['total'] or 0
    despesas_total = qs.aggregate(total=Sum('total'))['total'] or 0

    # 🔹 Acumulado mensal
    acumulado_por_mes = (
        qs.values('ano', 'mes')
          .annotate(soma=Sum('total'))
          .order_by('ano', 'mes')
    )

    acumulado_cumulativo = []
    running = 0
    for row in acumulado_por_mes:
        running += row['soma'] or 0
        acumulado_cumulativo.append({
            'ano': row['ano'],
            'mes': row['mes'],
            'soma': row['soma'],
            'acumulado': running
        })

    # 🔹 Paginação (AGORA no lugar certo)
    paginator = Paginator(qs, 6)
    page_number = request.GET.get('page')
    financeiros = paginator.get_page(page_number)

    return render(request, 'mvb/lista_financeiro.html', {
        'financeiros': financeiros,
        'receita_total': receita_total,
        'despesas_total': despesas_total,
        'acumulado_por_mes': acumulado_por_mes,
        'acumulado_cumulativo': acumulado_cumulativo,
    })


@login_required
def adicionar_financeiro(request):
    form = FinanceiroForm(request.POST or None)
    if form.is_valid():
        financeiro = form.save(commit=False)
        financeiro.criado_por = request.user
        financeiro.save()
        messages.success(request, f'Lançamento financeiro de {financeiro.data.strftime("%d/%m/%Y")} salvo com sucesso.')
        return redirect('lista_financeiro')
    return render(request, 'mvb/form.html', {'form': form, 'titulo': 'Adicionar Financeiro'})


@admin_required
def editar_financeiro(request, pk):
    financeiro = get_object_or_404(Financeiro, pk=pk)

    if request.method == "POST":
        form = FinanceiroForm(request.POST, instance=financeiro)
        if form.is_valid():
            form.save()
            return redirect("lista_financeiro")
    else:
        form = FinanceiroForm(instance=financeiro)

    return render(request, "mvb/editar_financeiro.html", {
        "form": form,
        "financeiro": financeiro
    })

@admin_required
def excluir_financeiro(request, pk):
    financeiro = get_object_or_404(Financeiro, pk=pk)

    if request.method == "POST":
        financeiro.delete()
        return redirect("lista_financeiro")

    return render(request, "mvb/excluir_financeiro.html", {
        "financeiro": financeiro
    })

@login_required
@permission_required('mvb.add_financeiro', raise_exception=True)
def novo_financeiro(request):
    if request.method == "POST":
        form = FinanceiroForm(request.POST)
        if form.is_valid():
            fin = form.save(commit=False)
            fin.criado_por = request.user
            fin.save()  # o save() já recalcula o total
            messages.success(request, "Lançamento financeiro salvo.")
            return redirect('lista_financeiro')
        else:
            messages.error(request, "Erro no formulário. Verifique os campos.")
    else:
        form = FinanceiroForm()

    return redirect('lista_financeiro')

@login_required
@permission_required('mvb.view_financeiro', raise_exception=True)
def exportar_financeiro_excel(request):
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')

    qs = Financeiro.objects.all().order_by('-ano', '-mes')
    if ano:
        qs = qs.filter(ano=int(ano))
    if mes:
        qs = qs.filter(mes=int(mes))

    rows = []
    total_geral = 0
    for f in qs:
        rows.append({
            "Data": f.data.strftime("%d/%m/%y") if f.data else "",
            "Ano": f.ano,
            "Mes": f.mes,
            "Salarios": float(f.salario_total_funcionarios or 0),
            "Frete": float(f.frete or 0),
            "Cafe": float(f.refeicao_cafe or 0),
            "Almoço": float(f.refeicao_almoco or 0),
            "Contabilidade": float(f.contabilidade or 0),
            "INSS": float(f.inss or 0),
            "Total (R$)": float(f.total or 0),
        })
        total_geral += float(f.total or 0)

    # linha final
    rows.append({
        "Data": "",
        "Ano": "",
        "Mes": "",
        "Salarios": "",
        "Frete": "",
        "Cafe": "",
        "Almoço": "",
        "Contabilidade": "",
        "INSS": "TOTAL GERAL",
        "Total (R$)": total_geral,
    })

    df = pd.DataFrame(rows)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=financeiro_relatorio.xlsx'
    df.to_excel(response, index=False)
    return response

def gerar_dados_financeiro():
    dados = []
    registros = Financeiro.objects.all().order_by('-data')

    for f in registros:
        dados.append({
            "Data": f.data.strftime("%d/%m/%Y"),
            "Ano": f.ano,
            "Mês": f.mes,
            "Salários": float(f.salario_total_funcionarios or 0),
            "Frete": float(f.frete or 0),
            "Café": float(f.refeicao_cafe or 0),
            "Almoço": float(f.refeicao_almoco or 0),
            "Contabilidade": float(f.contabilidade or 0),
            "INSS": float(f.inss or 0),
            "Total": float(f.total or 0),
        })

    return dados

@login_required
def financeiro_grafico_png(request):
    # agrupamento por mes
    qs = Financeiro.objects.values('ano','mes').annotate(total=Sum('total')).order_by('ano','mes')
    labels = [f"{r['ano']}-{r['mes']}" for r in qs]
    values = [float(r['total'] or 0) for r in qs]

    plt.figure(figsize=(6,2.5))
    plt.plot(labels, values, marker='o')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close()
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='image/png')

@login_required
@permission_required('mvb.view_financeiro', raise_exception=True)
def exportar_financeiro_pdf(request):
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')

    qs = Financeiro.objects.all().order_by('ano','mes')
    if ano: qs = qs.filter(ano=int(ano))
    if mes: qs = qs.filter(mes=int(mes))

    # montar tabela
    data_table = [["Data","Ano","Mês","Salários","Frete","Café","Almoço","Contab.","INSS","Total"]]
    total_geral = 0
    x_labels = []
    y_totals = []

    for f in qs:
        data_table.append([
            f.criado_em.strftime("%Y-%m-%d"),
            str(f.ano),
            str(f.mes),
            f"{f.salario_total_funcionarios:.2f}",
            f"{f.frete:.2f}",
            f"{f.refeicao_cafe:.2f}",
            f"{f.refeicao_almoco:.2f}",
            f"{f.contabilidade:.2f}",
            f"{f.inss:.2f}",
            f"{f.total:.2f}",
        ])
        total_geral += float(f.total or 0)
        x_labels.append(f"{f.ano}-{f.mes}")
        y_totals.append(float(f.total or 0))

    # linha total
    data_table.append(["","","","","","","","TOTAL GERAL","","",f"{total_geral:.2f}"])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Relatório Financeiro", styles['Title']), Spacer(1,12)]

    # tabela
    table = Table(data_table, repeatRows=1)
    table.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elems.append(table)
    elems.append(Spacer(1,12))
    
    # gráfico (barras)
    if x_labels and y_totals:
        plt.figure(figsize=(8,3))
        plt.bar(x_labels, y_totals)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        imgbuf = io.BytesIO()
        plt.savefig(imgbuf, format='png')
        plt.close()
        imgbuf.seek(0)
        im = Image(imgbuf, width=450, height=150)
        elems.append(Paragraph("Total por Período", styles['Heading3']))
        elems.append(im)

    doc.build(elems)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=financeiro_relatorio.pdf'
    return response

# Relatório (reaproveitável)
@login_required
def relatorio_periodo(request, periodo='diario'):
    hoje = date.today()
    if periodo == 'diario':
        dt_inicio = dt_fim = hoje
    elif periodo == 'semanal':
        dt_inicio = hoje - timedelta(days=hoje.weekday())
        dt_fim = dt_inicio + timedelta(days=6)
    elif periodo == 'mensal':
        dt_inicio = hoje.replace(day=1)
        if dt_inicio.month == 12:
            prox = dt_inicio.replace(year=dt_inicio.year+1, month=1, day=1)
        else:
            prox = dt_inicio.replace(month=dt_inicio.month+1, day=1)
        dt_fim = prox - timedelta(days=1)
    else:
        start = request.GET.get('start'); end = request.GET.get('end')
        if start and end:
            dt_inicio = date.fromisoformat(start)
            dt_fim = date.fromisoformat(end)
        else:
            dt_inicio = dt_fim = hoje

    receita_carretas = LavagemCarreta.objects.filter(data__range=(dt_inicio, dt_fim)).aggregate(
        total=Sum(ExpressionWrapper(F('quantidade_caixas') * F('valor_por_caixa'), output_field=DecimalField()))
    )['total'] or Decimal('0.00')

    receita_sujo = LavadorSujoEntry.objects.filter(data__range=(dt_inicio, dt_fim)).aggregate(
        total=Sum(ExpressionWrapper(F('quantidade_caixas') * F('valor_por_caixa'), output_field=DecimalField()))
    )['total'] or Decimal('0.00')

    receita_carga = LavadorCargaEntry.objects.filter(data__range=(dt_inicio, dt_fim)).aggregate(
        total=Sum('valor_rendido')
    )['total'] or Decimal('0.00')

    receita_total = receita_carretas + receita_sujo + receita_carga

    meses = set(); cursor = dt_inicio
    while cursor <= dt_fim:
        meses.add((cursor.year, cursor.month))
        cursor = cursor + timedelta(days=1)

    despesas_total = Decimal('0.00')
    financeiro_qs = Financeiro.objects.filter(Q(ano__in=[y for y,m in meses]) & Q(mes__in=[m for y,m in meses]))
    for f in financeiro_qs:
        despesas_total += f.total or Decimal("0.00")

    lucro = receita_total - despesas_total

    context = {
        'periodo': periodo,
        'dt_inicio': dt_inicio,
        'dt_fim': dt_fim,
        'receita_carretas': receita_carretas,
        'receita_sujo': receita_sujo,
        'receita_carga': receita_carga,
        'receita_total': receita_total,
        'despesas_total': despesas_total,
        'lucro': lucro,
    }
    return render(request, 'mvb/relatorio.html', context)

@login_required
def lista_clientes(request):
    cliente = Cliente.objects.all()
    return render(request, 'mvb/lista_clientes.html', {'clientes': cliente})

@admin_required
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect("lista_clientes")
    else:
        form = ClienteForm(instance=cliente)

    return render(request, "mvb/editar_cliente.html", {
        "form": form,
        "cliente": cliente
    })

@admin_required
def excluir_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == "POST":
        cliente.delete()
        return redirect("lista_clientes")

    return render(request, "mvb/excluir_cliente.html", {
        "cliente": cliente
    })


@admin_required
def novo_cliente(request):
    form = ClienteForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente criado com sucesso.")
            return redirect("lista_clientes")
        else:
            messages.error(request, "Erro ao salvar cliente.")
    return render(request, "mvb/cliente_form.html", {"form": form})

@login_required
def lista_presencas(request):
    funcionario_id = request.GET.get("func")

    presencas = Presenca.objects.select_related("funcionario").order_by("-data", "funcionario__nome")

    # Se filtrou por funcionário
    if funcionario_id:
        presencas = presencas.filter(funcionario_id=funcionario_id)

    funcionarios = Funcionario.objects.filter(ativo=True)

    return render(request, "mvb/lista_presencas.html", {
        "presencas": presencas,
        "funcionarios": funcionarios,
        "func_selected": funcionario_id,
    })

@login_required
def editar_presenca(request, pk):
    presenca = get_object_or_404(Presenca, pk=pk)

    if request.method == "POST":
        novo_status = request.POST.get("status")
        if novo_status in ["P", "F", "O"]:
            presenca.status = novo_status
            presenca.save()
            return redirect("lista_presencas")

    return render(request, "mvb/editar_presenca.html", {
        "presenca": presenca
    })

@login_required
def excluir_presenca(request, pk):
    presenca = get_object_or_404(Presenca, pk=pk)

    if request.method == "POST":
        presenca.delete()
        messages.success(request, "Presença removida com sucesso.")
        return redirect("lista_presencas")

    return render(request, "mvb/confirmar_excluir_presenca.html", {
        "presenca": presenca
    })

@login_required
def registrar_presencas(request):

    funcionarios = Funcionario.objects.all()
    data_selecionada = request.GET.get("data")

    if request.method == "POST":
        data = request.POST.get("data")
        if not data:
            messages.error(request, "Selecione uma data.")
            return redirect("registrar_presencas")

        # Salva cada funcionário
        for funcionario in funcionarios:
            status = request.POST.get(f"status_{funcionario.id}")  # P / F / O / vazio

            if status:
                Presenca.objects.update_or_create(
                    funcionario=funcionario,
                    data=data,
                    defaults={"status": status},
                )

        messages.success(request, "Presenças registradas com sucesso.")
        return redirect("lista_presencas")

    # Se usuário abriu a página com ?data=AAAA-MM-DD
    presencas_dict = {}
    if data_selecionada:
        presencas = Presenca.objects.filter(data=data_selecionada)
        presencas_dict = {p.funcionario.id: p.status for p in presencas}

    return render(request, "mvb/registrar_presencas.html", {
        "funcionarios": funcionarios,
        "data": data_selecionada,
        "presencas_dict": presencas_dict,
    })

def funcionario_sem_faltas_na_semana(funcionario, data_referencia):
    segunda = data_referencia - timedelta(days=data_referencia.weekday())
    domingo = segunda + timedelta(days=6)

    # Falta real = status F
    falta = Presenca.objects.filter(
        funcionario=funcionario,
        data__range=[segunda, domingo],
        status="F"
    ).exists()

    return not falta

def funcionario_sem_faltas_no_mes(funcionario, data_referencia):
    ano = data_referencia.year
    mes = data_referencia.month

    falta = Presenca.objects.filter(
        funcionario=funcionario,
        data__year=ano,
        data__month=mes,
        status="F"
    ).exists()

    return not falta

@login_required
def verificar_eligiveis_e_conceder(request):
    hoje = date.today()

    # Semana atual (segunda a domingo)
    inicio_sem = hoje - timedelta(days=hoje.weekday())
    fim_sem = inicio_sem + timedelta(days=6)

    # Mês atual
    inicio_mes = hoje.replace(day=1)
    _, last_day = monthrange(inicio_mes.year, inicio_mes.month)
    fim_mes = inicio_mes.replace(day=last_day)

    funcionarios = Funcionario.objects.filter(ativo=True)

    dados = []  # estrutura final enviada para o template

    for f in funcionarios:
        faltas_sem = Presenca.objects.filter(
            funcionario=f, data__range=(inicio_sem, fim_sem), status="F"
        ).count()

        faltas_mes = Presenca.objects.filter(
            funcionario=f, data__range=(inicio_mes, fim_mes), status="F"
        ).count()

        presencas_sem = Presenca.objects.filter(
            funcionario=f, data__range=(inicio_sem, fim_sem), status="P"
        ).count()

        folgas_sem = Presenca.objects.filter(
            funcionario=f, data__range=(inicio_sem, fim_sem), status="O"
        ).count()

        dados.append({
            "funcionario": f,
            "pres_sem": presencas_sem,
            "faltas_sem": faltas_sem,
            "folgas_sem": folgas_sem,
            "elegivel_sem": faltas_sem == 0,
            "elegivel_mes": faltas_mes == 0,
        })

    if request.method == "POST":
        with transaction.atomic():
            for item in dados:
                f = item["funcionario"]

                if request.POST.get(f"sem_{f.id}") == "on" and item["elegivel_sem"]:
                    BonusPayment.objects.create(
                        funcionario=f,
                        tipo="semanal",
                        valor=Decimal("120.00"),
                        criado_por=request.user,
                    )

                if request.POST.get(f"mes_{f.id}") == "on" and item["elegivel_mes"]:
                    BonusPayment.objects.create(
                        funcionario=f,
                        tipo="cesta",
                        valor=Decimal("0.00"),
                        criado_por=request.user,
                    )

        messages.success(request, "Bônus e cestas concedidos com sucesso!")
        return redirect("verificar_eligiveis_e_conceder")

    return render(request, "mvb/bonus_grant.html", {
        "dados": dados,
    })

# Exportar Relatório Excel (admin)
@admin_required
def export_relatorio_excel(request, periodo='mensal'):
    hoje = date.today()
    # definir dt_inicio/dt_fim como no relatorio_periodo
    if periodo == 'mensal':
        dt_inicio = hoje.replace(day=1)
        if dt_inicio.month == 12:
            prox = dt_inicio.replace(year=dt_inicio.year+1, month=1, day=1)
        else:
            prox = dt_inicio.replace(month=dt_inicio.month+1, day=1)
        dt_fim = prox - timedelta(days=1)
    else:
        dt_inicio = dt_fim = hoje

    lavagens = LavagemCarreta.objects.filter(data__range=(dt_inicio, dt_fim))
    sujos = LavadorSujoEntry.objects.filter(data__range=(dt_inicio, dt_fim))
    cargas = LavadorCargaEntry.objects.filter(data__range=(dt_inicio, dt_fim))
    financeiro_qs = Financeiro.objects.filter(ano__in=[dt_inicio.year, dt_fim.year], mes__in=[dt_inicio.month, dt_fim.month])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    rows = [
        ["Período", f"{dt_inicio} a {dt_fim}"],
        [],
        ["Receitas"],
        ["Tipo", "Quantidade", "Valor Total (R$)"],
    ]

    receita_carretas_total = sum([c.quantidade_caixas * c.valor_por_caixa for c in lavagens])
    rows.append(["Lavagens Carretas", sum([c.quantidade_caixas for c in lavagens]), float(receita_carretas_total)])
    receita_sujos_total = sum([s.quantidade_caixas * s.valor_por_caixa for s in sujos])
    rows.append(["Lavador Sujo", sum([s.quantidade_caixas for s in sujos]), float(receita_sujos_total)])
    receita_cargas_total = sum([c.valor_rendido for c in cargas])
    rows.append(["Lavador Carga (valor rendido)", sum([c.quantidade_caixas for c in cargas]), float(receita_cargas_total)])

    rows.append([])
    rows.append(["Despesas"])
    rows.append(["Ano","Mês","Salário Total","Frete","Café","Almoço","Contabilidade","INSS","Total Despesas"])
    for f in financeiro_qs:
        rows.append([f.ano, f.mes, float(f.salario_total_funcionarios), float(f.frete), float(f.refeicao_cafe), float(f.refeicao_almoco), float(f.contabilidade), float(f.inss), float(f.total_despesas())])

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=relatorio_{periodo}_{dt_inicio}_{dt_fim}.xlsx'
    wb.save(response)
    return response

# Exportar Relatório PDF (admin)
@admin_required
def export_relatorio_pdf(request, periodo='mensal'):
    hoje = date.today()
    if periodo == 'mensal':
        dt_inicio = hoje.replace(day=1)
        if dt_inicio.month == 12:
            prox = dt_inicio.replace(year=dt_inicio.year+1, month=1, day=1)
        else:
            prox = dt_inicio.replace(month=dt_inicio.month+1, day=1)
        dt_fim = prox - timedelta(days=1)
    else:
        dt_inicio = dt_fim = hoje

    lavagens = LavagemCarreta.objects.filter(data__range=(dt_inicio, dt_fim))
    sujos = LavadorSujoEntry.objects.filter(data__range=(dt_inicio, dt_fim))
    cargas = LavadorCargaEntry.objects.filter(data__range=(dt_inicio, dt_fim))
    financeiro_qs = Financeiro.objects.filter(ano__in=[dt_inicio.year, dt_fim.year], mes__in=[dt_inicio.month, dt_fim.month])

    context = {
        'dt_inicio': dt_inicio,
        'dt_fim': dt_fim,
        'lavagens': lavagens,
        'sujos': sujos,
        'cargas': cargas,
        'financeiro_qs': financeiro_qs,
    }
    html_string = render_to_string('mvb/relatorio_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_{periodo}_{dt_inicio}_{dt_fim}.pdf'
    return response

@admin_required
def exportar_excel_filtrado(request):
    tipo = request.GET.get("tipo")
    data = request.GET.get("data")
    cliente = request.GET.get("cliente")
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")

    incluir_lavagens = request.GET.get("incluir_lavagens")
    incluir_financeiro = request.GET.get("incluir_financeiro")
    incluir_presencas = request.GET.get("incluir_presencas")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=relatorio_filtrado.xlsx"

    with pd.ExcelWriter(response, engine="openpyxl") as writer:

        # ======================
        # LAVAGENS
        # ======================
        if incluir_lavagens:
            lavagens = filtrar_lavagens(
                tipo=tipo,
                data=data,
                cliente=cliente,
                data_inicio=data_inicio,
                data_fim=data_fim
            )

            dados_lavagens = []
            total_geral = 0

            for l in lavagens:
                valor_unitario = getattr(l, "valor_por_caixa", None) or getattr(l, "valor_rendido", 0)
                valor_unitario = float(valor_unitario or 0)

                quantidade = l.quantidade_caixas or 0
                subtotal = quantidade * valor_unitario
                total_geral += subtotal

                dados_lavagens.append({
                    "Data": l.data.strftime("%d/%m/%Y"),
                    "Cliente": l.cliente.nome if l.cliente else "",
                    "Tipo": l.get_tipo_lavagem_display(),
                    "Produto": getattr(l, "tipo_produto", ""),
                    "Tipo Caixa": getattr(l, "tipo_caixa", getattr(l, "tamanho_caixa", "")),
                    "Quantidade": quantidade,
                    "Valor Unitário": valor_unitario,
                    "Subtotal": subtotal,
                })

            df_lavagens = pd.DataFrame(dados_lavagens)

            # ➕ Linha TOTAL (forma segura)
            if not df_lavagens.empty:
                linha_total_lavagens = {col: "" for col in df_lavagens.columns}
                linha_total_lavagens["Subtotal"] = total_geral

                df_lavagens = pd.concat(
                    [df_lavagens, pd.DataFrame([linha_total_lavagens])],
                    ignore_index=True
                )

            df_lavagens.to_excel(writer, sheet_name="Lavagens", index=False)

        # ======================
        # FINANCEIRO
        # ======================
        if incluir_financeiro:
            financeiros = Financeiro.objects.all()

            if data:
                financeiros = financeiros.filter(data=data)
            if data_inicio:
                financeiros = financeiros.filter(data__gte=data_inicio)
            if data_fim:
                financeiros = financeiros.filter(data__lte=data_fim)

            dados_financeiro = []
            total_financeiro = 0

            for f in financeiros:
                total_financeiro += float(f.total)

                dados_financeiro.append({
                    "Data": f.data.strftime("%d/%m/%Y"),
                    "Salários": float(f.salario_total_funcionarios),
                    "Frete": float(f.frete),
                    "Café": float(f.refeicao_cafe),
                    "Almoço": float(f.refeicao_almoco),
                    "Contabilidade": float(f.contabilidade),
                    "INSS": float(f.inss),
                    "Total": float(f.total),
                })

            df_fin = pd.DataFrame(dados_financeiro)

            # ➕ Linha TOTAL (forma segura)
            if not df_fin.empty:
                linha_total_fin = {col: "" for col in df_fin.columns}
                linha_total_fin["Total"] = total_financeiro

                df_fin = pd.concat(
                    [df_fin, pd.DataFrame([linha_total_fin])],
                    ignore_index=True
                )

            df_fin.to_excel(writer, sheet_name="Financeiro", index=False)

        # ======================
        # PRESENÇAS
        # ======================
        if incluir_presencas:
            presencas = filtrar_presencas(
                funcionario=cliente,  # se você reutilizar o select
                data=data,
                data_inicio=data_inicio,
                data_fim=data_fim
            )

            dados_presencas = []

            total_presentes = 0
            total_faltas = 0

            for p in presencas:
                status = p.get_status_display()

                if status.lower() == "presente":
                    total_presentes += 1
                elif status.lower() == "falta":
                    total_faltas += 1

                dados_presencas.append({
                    "Data": p.data.strftime("%d/%m/%Y"),
                    "Funcionário": p.funcionario.nome,
                    "Status": status,
                })

            df_presencas = pd.DataFrame(dados_presencas)

            if not df_presencas.empty:
                linha_total = {
                    "Data": "",
                    "Funcionário": "TOTAL",
                    "Status": f"Presentes: {total_presentes} | Faltas: {total_faltas}"
                }

                df_presencas = pd.concat(
                    [df_presencas, pd.DataFrame([linha_total])],
                    ignore_index=True
                )

            df_presencas.to_excel(writer, sheet_name="Presenças", index=False)

    return response

@admin_required
def exportar_pdf_filtrado(request):
    tipo = request.GET.get("tipo")
    data = request.GET.get("data")
    cliente = request.GET.get("cliente")

    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")

    incluir_lavagens = request.GET.get("incluir_lavagens")
    incluir_financeiro = request.GET.get("incluir_financeiro")
    incluir_presencas = request.GET.get("incluir_presencas")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=relatorio_filtrado.pdf"

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()

    elementos = [Paragraph("Relatório Filtrado", styles["Title"])]

    # =========================
    # LAVAGENS (SÓ SE MARCADO)
    # =========================
    if incluir_lavagens:
        lavagens = filtrar_lavagens(
            tipo=tipo,
            data=data,
            cliente=cliente,
            data_inicio=data_inicio,
            data_fim=data_fim,
        )

        data_table = [
            ["Data", "Cliente", "Tipo", "Produto", "Tipo Caixa", "Qtd", "Valor Unit.", "Total"]
        ]

        total_geral = 0

        for l in lavagens:
            valor_unitario = getattr(l, "valor_por_caixa", None) or getattr(l, "valor_rendido", 0) or 0
            quantidade = l.quantidade_caixas or 0
            subtotal = quantidade * float(valor_unitario)

            total_geral += subtotal

            data_table.append([
                l.data.strftime("%d/%m/%Y"),
                l.cliente.nome if l.cliente else "",
                l.get_tipo_lavagem_display(),
                getattr(l, "tipo_produto", ""),
                getattr(l, "tipo_caixa", getattr(l, "tamanho_caixa", "")),
                quantidade,
                f"R$ {valor_unitario:.2f}",
                f"R$ {subtotal:.2f}",
            ])

        data_table.append(["", "", "", "", "", "", "TOTAL", f"R$ {total_geral:.2f}"])

        table = Table(data_table, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME", (-2,-1), (-1,-1), "Helvetica-Bold"),
            ("BACKGROUND", (-2,-1), (-1,-1), colors.beige),
        ]))

        elementos.append(Paragraph("Lavagens", styles["Heading2"]))
        elementos.append(table)

    # =========================
    # FINANCEIRO (SÓ SE MARCADO)
    # =========================
    if incluir_financeiro:
        financeiros = Financeiro.objects.all()

        if data:
            financeiros = financeiros.filter(data=data)

        if data_inicio:
            financeiros = financeiros.filter(data__gte=data_inicio)

        if data_fim:
            financeiros = financeiros.filter(data__lte=data_fim)

        tabela_financeiro = [
            ["Data", "Salários", "Frete", "Café", "Almoço", "Contabilidade", "INSS", "Total"]
        ]

        total_geral_financeiro = 0

        for f in financeiros:
            total_geral_financeiro += float(f.total)

            tabela_financeiro.append([
                f.data.strftime("%d/%m/%Y"),
                f"R$ {f.salario_total_funcionarios:.2f}",
                f"R$ {f.frete:.2f}",
                f"R$ {f.refeicao_cafe:.2f}",
                f"R$ {f.refeicao_almoco:.2f}",
                f"R$ {f.contabilidade:.2f}",
                f"R$ {f.inss:.2f}",
                f"R$ {f.total:.2f}",
            ])

        tabela_financeiro.append([
            "", "", "", "", "", "", "TOTAL", f"R$ {total_geral_financeiro:.2f}"
        ])

        table_fin = Table(tabela_financeiro, repeatRows=1)
        table_fin.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME", (-2,-1), (-1,-1), "Helvetica-Bold"),
            ("BACKGROUND", (-2,-1), (-1,-1), colors.beige),
        ]))

        elementos.append(Paragraph("Financeiro", styles["Heading2"]))
        elementos.append(table_fin) 

    # ======================
    # PRESENÇAS (PDF)
    # ======================
    if incluir_presencas:
        presencas = filtrar_presencas(
            funcionario=cliente,
            data=data,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        tabela_presencas = [
            ["Data", "Funcionário", "Status"]
        ]

        total_presentes = 0
        total_faltas = 0

        for p in presencas:
            status = p.get_status_display()

            if status.lower() == "presente":
                total_presentes += 1
            elif status.lower() == "falta":
                total_faltas += 1

            tabela_presencas.append([
                p.data.strftime("%d/%m/%Y"),
                p.funcionario.nome,
                status,
            ])

        # Linha de totais
        tabela_presencas.append([
            "",
            "TOTAL",
            f"Presentes: {total_presentes} | Faltas: {total_faltas}"
        ])

        table = Table(tabela_presencas, colWidths=[80, 220, 200])

        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTNAME", (1,-1), (-1,-1), "Helvetica-Bold"),
            ("BACKGROUND", (1,-1), (-1,-1), colors.beige),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("ALIGN", (1,1), (1,-1), "LEFT"),
        ]))

        elementos.append(Paragraph("Relatório de Presenças", styles["Heading2"]))
        elementos.append(table)   

    doc.build(elementos)
    return response

# View da tela inicial (formulário)
@admin_required
def relatorios(request):
    from .models import Cliente
    clientes = Cliente.objects.all()
    return render(request, "mvb/relatorios.html", {"clientes": clientes})

# View que processa o relatório
@admin_required
def resultado_relatorio(request):
    # filtros comuns
    tipo = request.GET.get("tipo")
    data = request.GET.get("data")
    cliente = request.GET.get("cliente")

    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")

    incluir_lavagens = request.GET.get("incluir_lavagens")
    incluir_financeiro = request.GET.get("incluir_financeiro")
    incluir_presencas = request.GET.get("incluir_presencas")

    # =========================
    # LAVAGENS
    # =========================
    lavagens_processadas = []
    total_geral = Decimal("0.00")

    if incluir_lavagens:
        lavagens_queryset = filtrar_lavagens(
            tipo=tipo,
            data=data,
            cliente=cliente,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        for l in lavagens_queryset:
            valor_unitario = getattr(l, "valor_por_caixa", None) or getattr(l, "valor_rendido", 0)

            try:
                valor_unitario = Decimal(valor_unitario)
            except:
                valor_unitario = Decimal("0.00")

            quantidade = l.quantidade_caixas or 0
            subtotal = quantidade * valor_unitario
            total_geral += subtotal

            lavagens_processadas.append({
                "data": l.data,
                "cliente": l.cliente.nome if l.cliente else "",
                "tipo_lavagem": l.get_tipo_lavagem_display(),
                "produto": getattr(l, "tipo_produto", ""),
                "tipo_caixa": getattr(l, "tipo_caixa", getattr(l, "tamanho_caixa", "")),
                "quantidade": quantidade,
                "valor_unitario": valor_unitario,
                "subtotal": subtotal,
            })

    # =========================
    # FINANCEIRO
    # =========================
    financeiros_queryset = []
    total_financeiro = Decimal("0.00")

    if incluir_financeiro:
        financeiros_queryset = Financeiro.objects.all()

        if data:
            financeiros_queryset = financeiros_queryset.filter(data=data)

        if data_inicio:
            financeiros_queryset = financeiros_queryset.filter(data__gte=data_inicio)

        if data_fim:
            financeiros_queryset = financeiros_queryset.filter(data__lte=data_fim)

        total_financeiro = financeiros_queryset.aggregate(
            total=Sum("total")
        )["total"] or Decimal("0.00")

    # =========================
    # PRESENÇAS
    # =========================
    presencas_processadas = []

    if incluir_presencas:
        presencas = Presenca.objects.select_related("funcionario").all()

        if data:
            presencas = presencas.filter(data=data)

        if data_inicio:
            presencas = presencas.filter(data__gte=data_inicio)

        if data_fim:
            presencas = presencas.filter(data__lte=data_fim)

        if cliente:
            presencas = presencas.filter(funcionario__nome__icontains=cliente)

        for p in presencas.order_by("data", "funcionario__nome"):
            presencas_processadas.append({
                "data": p.data,
                "funcionario": p.funcionario.nome,
                "status": p.get_status_display(),
            })

    # =========================
    # RENDER
    # =========================
    return render(request, "mvb/resultado_relatorio.html", {
        "lavagens": lavagens_processadas,
        "financeiros": financeiros_queryset,
        "presencas": presencas_processadas,

        "incluir_lavagens": incluir_lavagens,
        "incluir_financeiro": incluir_financeiro,
        "incluir_presencas": incluir_presencas,

        "tipo": tipo,
        "data": data,
        "cliente": cliente,
        "data_inicio": data_inicio,
        "data_fim": data_fim,

        "total_geral": total_geral,
        "total_financeiro": total_financeiro,
    })

def filtrar_lavagens(tipo=None, data=None, cliente=None, data_inicio=None, data_fim=None):
    lavagens = LavagemCarreta.objects.all()

    # FILTRO POR TIPO
    if tipo:
        lavagens = lavagens.filter(tipo_lavagem=tipo)

    # FILTRO POR CLIENTE
    if cliente:
        lavagens = lavagens.filter(cliente__nome__icontains=cliente)

    # FILTRO POR DATA EXATA
    if data:
        try:
            d = datetime.strptime(data, "%Y-%m-%d").date()
            lavagens = lavagens.filter(data=d)
        except:
            pass

    # FILTRO POR PERÍODO
    if data_inicio and data_fim:
        try:
            di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            df = datetime.strptime(data_fim, "%Y-%m-%d").date()
            lavagens = lavagens.filter(data__range=[di, df])
        except:
            pass

    return lavagens

def filtrar_presencas(funcionario=None, data=None, data_inicio=None, data_fim=None):
    qs = Presenca.objects.select_related("funcionario")

    if funcionario:
        qs = qs.filter(funcionario__nome=funcionario)

    if data:
        qs = qs.filter(data=data)

    if data_inicio:
        qs = qs.filter(data__gte=data_inicio)

    if data_fim:
        qs = qs.filter(data__lte=data_fim)

    return qs.order_by("data", "funcionario__nome")

@admin_required
def relatorio_por_cliente(request, cliente_id=None):
    if cliente_id:
        cliente = get_object_or_404(Cliente, pk=cliente_id)
        lavagens = LavagemCarreta.objects.filter(cliente=cliente)
        sujos = LavadorSujoEntry.objects.filter(cliente=cliente)
        cargas = LavadorCargaEntry.objects.filter(cliente=cliente)
    else:
        cliente = None
        lavagens = sujos = cargas = []
    return render(request, 'mvb/relatorio_cliente.html', {'cliente': cliente, 'lavagens': lavagens, 'sujos': sujos, 'cargas': cargas})

@login_required
def exportar_lavagens_excel(request):
    tipo = request.GET.get("tipo")
    data = request.GET.get("data")
    cliente = request.GET.get("cliente")

    lavagens = filtrar_lavagens(tipo, data, cliente)

    linhas = []
    total_geral = 0

    for l in lavagens:
        valor_unitario = getattr(l, "valor_por_caixa", None) or getattr(l, "valor_rendido", None)
        subtotal = l.quantidade_caixas * valor_unitario if valor_unitario else 0

        total_geral += subtotal

        linhas.append({
            "Data": l.data,
            "Cliente": l.cliente.nome if l.cliente else "",
            "Tipo Lavagem": l.get_tipo_lavagem_display(),
            "Produto": getattr(l, "tipo_produto", ""),
            "Tipo Caixa": getattr(l, "tipo_caixa", getattr(l, "tamanho_caixa", "")),
            "Quantidade": l.quantidade_caixas,
            "Valor Unitário (R$)": float(valor_unitario) if valor_unitario else 0,
            "Subtotal (R$)": float(subtotal),
        })

    # ADICIONAR LINHA EXTRA DO TOTAL GERAL
    linhas.append({
        "Data": "",
        "Cliente": "",
        "Tipo Lavagem": "",
        "Produto": "",
        "Tipo Caixa": "",
        "Quantidade": "",
        "Valor Unitário (R$)": "",
        "Subtotal (R$)": float(total_geral),
    })

    df = pd.DataFrame(linhas)

    # remover timezone do Data caso necessário
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="lavagens_filtradas.xlsx"'
    df.to_excel(response, index=False)

    return response

@login_required
def exportar_lavagens_pdf(request):
    tipo = request.GET.get("tipo")
    data = request.GET.get("data")
    cliente = request.GET.get("cliente")

    lavagens = filtrar_lavagens(tipo, data, cliente)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=lavagens_filtradas.pdf"

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()

    tabela = [
        ["Data", "Cliente", "Tipo", "Produto", "Caixa", "Qtd", "Valor Unit.", "Subtotal"],
    ]

    total_geral = 0

    for l in lavagens:
        valor_unitario = getattr(l, "valor_por_caixa", None) or getattr(l, "valor_rendido", None)
        subtotal = l.quantidade_caixas * valor_unitario if valor_unitario else 0
        total_geral += subtotal

        tabela.append([
            str(l.data),
            l.cliente.nome if l.cliente else "",
            l.get_tipo_lavagem_display(),
            getattr(l, "tipo_produto", ""),
            getattr(l, "tipo_caixa", getattr(l, "tamanho_caixa", "")),
            l.quantidade_caixas,
            f"R$ {valor_unitario:.2f}" if valor_unitario else "-",
            f"R$ {subtotal:.2f}",
        ])

    # ADICIONAR TOTAL GERAL
    tabela.append([
        "",
        "",
        "",
        "",
        "",
        "",
        "TOTAL GERAL:",
        f"R$ {total_geral:.2f}",
    ])

    table = Table(tabela)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("ALIGN", (5, 1), (7, -1), "RIGHT"),
        ("BACKGROUND", (6, -1), (7, -1), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))

    doc.build([
        Paragraph("Relatório de Lavagens", styles["Title"]),
        table
    ])

    return response